import openpyxl

path = r'F:\Scoopons\selinium_work\selenium_codes\datadriventesting\testdata.xlsx'

workbook = openpyxl.load_workbook(path)  # loading workbook

# get the sheet from workbook
sheet = workbook.active  # acting is nothing capture active_sheet and store into variable

# sheet=workbook.get_sheet_by_name("Sheet1")

rowcount = sheet.max_row
colcount = sheet.max_column

print("Number of rows: ", rowcount)
print("Number of columns: ", colcount)

for row in range(1, rowcount + 1):
    for col in range(1, colcount + 1):
        print(sheet.cell(row=row, column=col).value, end="  ")
    print()
